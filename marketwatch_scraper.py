import openpyxl as xl
import requests
from bs4 import BeautifulSoup


# code getFunctions for all individual attributes

def getCurrentPPS(soup):
    price_div = soup.find('h3', class_ = 'intraday__price')
    curr_price = price_div.find('bg-quote').contents[0]
    curr_price = float(curr_price.replace(',', ''))
    return curr_price


def getPercentChange(soup):
    intraday_div = soup.find('div', class_ = 'intraday__data')
    per_change_span = intraday_div.find('span', class_ = 'change--percent--q')
    per_change = per_change_span.find('bg-quote')
    try:
        if per_change.attrs['session'] == 'after':
            close_div = soup.find('div', class_ = 'intraday__close')
            per_change = close_div.find_all('td')[2].contents[0]
        else:
            per_change = per_change.contents[0]
    except KeyError:
        per_change = per_change.contents[0]

    per_change = float(per_change.replace(',', '')[:-1]) / 100
    return per_change


def getSector(soup):
    sector_div = soup.find('div', class_ = 'intraday__sector')
    sector = str(sector_div.find('span', class_ = 'label').contents[0])
    return sector


def getSectorChange(soup):
    sector_div = soup.find('div', class_ = 'intraday__sector')
    sector_change_div = sector_div.find('span', class_ = 'change--percent positive')
    if sector_change_div == None:
        sector_change_div = sector_div.find('span', class_ = 'change--percent negative')
    sector_change = sector_change_div.find('span').contents[0][:-1]
    sector_change = float(sector_change.replace(',', '')) / 100
    return sector_change


def getPerformance(soup):
    performance_div = soup.find('div', class_ = 'element element--table performance')
    performance = performance_div.find_all('li', class_ = 'content__item value ignore-color')
    performance = [float(perf.contents[0].replace('%', '')) / 100 for perf in performance]
    return performance


def getPrimaryAttributes(soup):
    primary_attributes = {}
    primary_div = soup.find('div', class_ = 'content-region region--primary')
    primary_items = primary_div.find_all('li', class_ = 'kv__item')
    for item in primary_items:
        key = str(item.find('small').contents[0])
        value = str(item.find('span').contents[0])
        primary_attributes[key] = value
    return primary_attributes


def getStockAttributes(ticker):
    # get the source code, then call individual get functions for each attribute, passing the source code to each one
    url = 'https://www.marketwatch.com/investing/stock/' + ticker.lower()
    page = requests.get(url, 'html.parser')
    if page.status_code != 200:
        raise Exception('Failed to retrieve URL contents')
    else:
        attributes = {}
        soup = BeautifulSoup(page.content, 'html.parser')

        attributes['Current PPS'] = getCurrentPPS(soup)
        attributes['% Change'] = getPercentChange(soup)
        attributes['Sector'] = getSector(soup)
        attributes['Sector % Change'] = getSectorChange(soup)
        performance = getPerformance(soup)
        attributes['5 Day Performance'] = performance[0]
        attributes['1 Month Performance'] = performance[1]
        attributes['3 Month Performance'] = performance[2]
        attributes['1 Year Performance'] = performance[3]
        attributes['YTD Performance'] = performance[4]
        primary_attributes = getPrimaryAttributes(soup)
        attributes['52 Week Low'] = float(primary_attributes['52 Week Range'].split(' - ')[0].replace(',', ''))
        attributes['52 Week High'] = float(primary_attributes['52 Week Range'].split(' - ')[1].replace(',', ''))
        try:
            attributes['P/E Ratio'] = float(primary_attributes['P/E Ratio'].replace(',', ''))
        except:
            attributes['P/E Ratio'] = '-'
        attributes['EPS'] = float(primary_attributes['EPS'].replace('$', ''))
        attributes['Market Cap'] = primary_attributes['Market Cap'][1:].replace(',', '')
        attributes['% of Float Shorted'] = float(primary_attributes['% of Float Shorted'].replace('%', '')) / 100
        return attributes


wb_path = '/Users/billy/Personal/Finances/Financials.xlsx'
wb = xl.load_workbook(wb_path)
sheet = wb['Stocks']
ticker_col = 1
first_ticker = sheet.min_row + 1
last_ticker = sheet.max_row
attributes_cell_keys = {'Current PPS': 'E',
                        '% Change': 'F',
                        'Sector': 'G',
                        'Sector % Change': 'H',
                        '5 Day Performance': 'I',
                        '1 Month Performance': 'J',
                        '3 Month Performance': 'K',
                        'YTD Performance': 'L',
                        '1 Year Performance': 'M',
                        '52 Week Low': 'N',
                        '52 Week High': 'O',
                        'P/E Ratio': 'P',
                        'EPS': 'Q',
                        'Market Cap': 'R',
                        '% of Float Shorted': 'S'}

for r in range(first_ticker, last_ticker + 1):
    ticker = sheet.cell(row = r, column = ticker_col).value

    try:
        attributes = getStockAttributes(ticker)
    except AttributeError:
        break
    print(ticker)
    for key in attributes.keys():
        value = attributes[key]
        column = attributes_cell_keys[key]
        cell = column + str(r)
        sheet[cell] = value
wb.save(filename = wb_path)
